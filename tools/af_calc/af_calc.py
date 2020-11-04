import openpyxl as xl
import sys
import numpy as np
from matplotlib import pyplot as plt
from PySide2.QtWidgets import QFileDialog
from tools.af_calc.af_calc_ui import Ui_AFCalcToolView
from ui.customwidget import MatplotlibWidget, SubWindow


class AfCalcTool(SubWindow):
    def __init__(self, name, parent=None):
        super().__init__(name, parent, Ui_AFCalcToolView())
        save_params = self.load_params()
        self.ui.okbutton.clicked.connect(self.calc_ret)
        self.ui.open_in_path.clicked.connect(self.open_in_path)
        self.ui.open_out_path.clicked.connect(self.open_out_path)
        self.matplot = MatplotlibWidget(self.ui.plotview)

    def open_in_path(self):
        filepath = QFileDialog.getOpenFileName(
            None, '打开excel', './', "Excel (*.xlsx *.xls)")
        if(filepath != ''):
            self.input_filepath = filepath[0]

    def open_out_path(self):
        filepath = QFileDialog.getSaveFileName(
            None, "选择保存文件", './', "txt (*.txt)")
        if(filepath != ''):
            self.output_filepath = filepath[0]

    def calc_ret(self):
        # 加载参数
        wb = xl.load_workbook(self.input_filepath)
        sheet = self.ui.sheetname.text()
        yr = self.ui.zoom_num.value()
        xr = self.ui.focus_num.value()
        z3 = self.ui.zoom_range.value()
        zp = self.ui.zoom_ratio.text()
        f4 = self.ui.tele_pianyi.value()
        fp = self.ui.focus_ratio.text()
        m = self.ui.focus_nums.text()
        m = list(m.strip().split(','))
        m = list(map(lambda x: int(x), m))
        start = self.ui.start_num.value()
        end = self.ui.end_num.value()
        skip = self.ui.num_inter.value()

        ws = wb[sheet]

        if zp == "":
            zp = 4
        else:
            zp = int(zp)

        if fp == "":
            fp = 2
        else:
            fp = int(fp)

        z3 *= zp
        f4 *= fp

        y = 0
        x = 0
        for i in list(ws.columns)[yr]:
            if i.value != 0:
                y += 1
            else:
                break
        for i in list(ws.columns)[xr]:
            if i.value != 0:
                x += 1
            else:
                break

        if end == "":
            end = min(x, y)+1
        else:
            end = int(end)

        rst = []
        for i in range(start, end+1):
            lst = []
            if (i-start) % skip != 0:
                continue
            lst.append(round(z3-ws.cell(i, yr).value*zp))
            for e in m:
                lst.append(round(f4+ws.cell(i, e).value*fp))
            rst.append(lst)

        # 画图
        for i in range(1, len(m)):
            self.matplot.input([l[0] for l in rst], [l[i] for l in rst])
        self.matplot.draw()

        # 写入文件
        s = "{"
        for i in rst:
            s += "{"
            for j in i:
                s += " "
                s += str(j)
                s += ","
            s = s[:-1]
            s += "},\n"
        s = s[:-2]
        s += "}"
        # print(s)
        f = open(self.output_filepath, 'w')
        f.write(s)
        f.close()
        return
